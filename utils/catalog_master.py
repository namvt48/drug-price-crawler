"""Nạp catalog chuẩn từ output/catalog_master.xlsx.

Ví von: trước đây app tự đi "kiểm kê" (crawl) danh mục từng site rồi tự đoán
sản phẩm nào trùng sản phẩm nào (fuzzy-match). File này là "sổ kiểm kê đã
được dược sĩ duyệt" làm sẵn ở ngoài (entity-resolution) — app chỉ cần đọc vào,
không tự đoán nữa. Sheet `master_products` cho tên chuẩn mỗi nhóm sản phẩm,
sheet `source_listings` cho biết nhóm đó có mặt ở site nào với product_id gì
(dùng để fetch giá live — file này KHÔNG có giá).
"""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook

from utils.config_loader import app_base_dir
from utils.models import CatalogItem, SourceName
from utils.normalizer import strip_accents
from utils.url_detect import detect_product_id

LogFn = Callable[[str], None]

_MASTER_SHEET = "master_products"
_LISTINGS_SHEET = "source_listings"
_SITE_ID_BY_SOURCE = {
    SourceName.BACHHOATHUOC: "bachhoathuoc",
    SourceName.CHOTHUOC247: "chothuoc247",
    SourceName.CHOTHUOCTOT: "chothuoctot",
    SourceName.DUOCPHAMGIASI: "duocphamgiasi",
    SourceName.GIATHUOCTOT: "giathuoctot",
    SourceName.THUOCHAPU: "thuochapu",
    SourceName.THUOCSI: "thuocsi",
    SourceName.THUOCSISAIGON: "thuocsisaigon",
    SourceName.THUOCTOT3MIEN: "thuoctot3mien",
}


def _default_path() -> Path:
    return app_base_dir() / "output" / "catalog_master.xlsx"


def load_master_catalog(
    path: str | Path | None = None, log: LogFn | None = None
) -> list[CatalogItem]:
    """Đọc `master_products` + `source_listings` → list[CatalogItem] phẳng,
    mỗi item mang sẵn `master_product_id` để `gui.viewmodel.build_catalog_groups`
    gộp nhóm theo đúng entity-resolution (không fuzzy-match lại). Thiếu file
    hoặc lỗi đọc → log rồi trả `[]`, không làm crash app."""
    log = log or (lambda _m: None)
    xlsx_path = Path(path) if path else _default_path()
    if not xlsx_path.exists():
        log(f"Không tìm thấy catalog chuẩn {xlsx_path} — catalog sẽ rỗng.")
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        log(f"Lỗi mở catalog chuẩn {xlsx_path}: {exc}")
        return []

    try:
        canonical_names = _load_canonical_names(wb, log)
        items = _load_listings(wb, canonical_names, log)
        items.sort(
            key=lambda item: (
                item.search_name,
                item.master_product_id,
                item.source.value,
            )
        )
        return items
    finally:
        wb.close()


def _load_canonical_names(wb, log: LogFn) -> dict[str, str]:
    if _MASTER_SHEET not in wb.sheetnames:
        log(f"Catalog chuẩn thiếu sheet '{_MASTER_SHEET}'.")
        return {}
    ws = wb[_MASTER_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}
    id_idx = header.index("master_product_id")
    name_idx = header.index("tên_sản_phẩm_chuẩn")
    names: dict[str, str] = {}
    for row in rows:
        master_id = row[id_idx]
        name = row[name_idx]
        if master_id and name:
            names[str(master_id)] = str(name)
    return names


def _load_listings(
    wb, canonical_names: dict[str, str], log: LogFn
) -> list[CatalogItem]:
    if _LISTINGS_SHEET not in wb.sheetnames:
        log(f"Catalog chuẩn thiếu sheet '{_LISTINGS_SHEET}'.")
        return []
    ws = wb[_LISTINGS_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    idx = {col: i for i, col in enumerate(header)}

    items: list[CatalogItem] = []
    skipped = 0
    valid_rows: list[tuple[tuple, str, SourceName]] = []
    pair_counts: dict[tuple[str, SourceName], int] = {}
    for row in rows:
        master_id = row[idx["master_product_id"]]
        source_raw = row[idx["source"]]
        product_id = row[idx["product_id"]]
        source_url = row[idx["source_url"]] if "source_url" in idx else None
        if not master_id or not source_raw or not product_id or not source_url:
            skipped += 1
            continue
        try:
            source = SourceName(source_raw)
        except ValueError:
            skipped += 1
            continue

        detected_id = detect_product_id(_SITE_ID_BY_SOURCE[source], str(source_url))
        if detected_id != str(product_id):
            skipped += 1
            log(
                "Catalog chuẩn: link không khớp product_id "
                f"{master_id}/{source.value}: {source_url} != {product_id}."
            )
            continue

        master_id = str(master_id)
        pair = (master_id, source)
        pair_counts[pair] = pair_counts.get(pair, 0) + 1
        valid_rows.append((row, master_id, source))

    duplicate_pairs = {pair for pair, count in pair_counts.items() if count > 1}
    for row, master_id, source in valid_rows:
        if (master_id, source) in duplicate_pairs:
            continue
        product_id = row[idx["product_id"]]
        canonical_name = canonical_names.get(master_id) or str(
            row[idx["drug_name"]] or ""
        )
        source_drug_name = str(row[idx["drug_name"]] or "")
        manufacturer = (
            row[idx["nhà_sản_xuất_xuất_xứ"]] if "nhà_sản_xuất_xuất_xứ" in idx else None
        )
        source_url = row[idx["source_url"]] if "source_url" in idx else None

        items.append(
            CatalogItem(
                product_id=str(product_id),
                drug_name=canonical_name,
                search_name=strip_accents(canonical_name).lower(),
                source_drug_name=source_drug_name,
                manufacturer=str(manufacturer) if manufacturer else "",
                source=source,
                source_url=str(source_url) if source_url else "",
                master_product_id=master_id,
            )
        )

    if skipped:
        log(f"Catalog chuẩn: bỏ qua {skipped} listing thiếu id/nguồn/product_id.")
    if duplicate_pairs:
        for master_id, source in sorted(
            duplicate_pairs, key=lambda pair: (pair[0], pair[1].value)
        ):
            log(
                "Catalog chuẩn: trùng master/site "
                f"{master_id}/{source.value} ({pair_counts[(master_id, source)]} dòng) "
                "— không nạp dòng nào vì không được đoán product_id."
            )
        log(
            "Catalog chuẩn: phát hiện "
            f"{len(duplicate_pairs)} cặp master/site bị trùng; "
            "toàn bộ cặp trùng được đánh lỗi."
        )
    return items


def _next_id(ws, id_column: str, prefix: str, width: int) -> str:
    """ID tiếp theo cho `id_column`, đọc TRỰC TIẾP dòng cuối (`ws.max_row`, O(1))
    thay vì quét cả sheet — quan trọng vì sheet có hàng chục nghìn dòng. Giữ nguyên
    độ rộng zero-pad hiện có (`MP` 6 số, `L` 7 số)."""
    header = [c.value for c in ws[1]]
    idx = header.index(id_column) + 1  # ws.cell dùng chỉ số 1-based
    last_value = ws.cell(row=ws.max_row, column=idx).value if ws.max_row > 1 else None
    if not last_value:
        return f"{prefix}{1:0{width}d}"
    num_part = str(last_value)[len(prefix) :]
    return f"{prefix}{int(num_part) + 1:0{width}d}"


def append_manual_product(
    items: list[CatalogItem],
    canonical_name: str,
    path: str | Path | None = None,
) -> str:
    """Ghi 1 sản phẩm THÊM TAY (tính năng 'Thêm sản phẩm mới' trong GUI — xem
    `gui.main_window._save_manual_product`) vào catalog: 1 dòng `master_products` +
    N dòng `source_listings` (1 dòng/site trong `items`). CHỈ ghi 2 sheet này —
    `site_id_lookup`/`product_variants`/`match_review`/`match_summary` không được
    `load_master_catalog` đọc, không đụng vào để giảm rủi ro hỏng file.

    Mở file ở chế độ ĐẦY ĐỦ (`read_only=False`) — CHẬM HƠN NHIỀU so với đọc (đo thật
    ~2-3 phút với file ~40k dòng sản xuất) vì phải tải nguyên object model để ghi
    được. PHẢI gọi trong thread nền, không gọi trực tiếp từ UI thread (xem
    `_save_manual_product`). Ghi qua file tạm
    rồi rename để không hỏng file gốc (41k+ dòng, khó khôi phục) nếu process bị tắt
    giữa lúc đang lưu."""
    sources = [item.source for item in items]
    if len(sources) != len(set(sources)):
        raise ValueError("Mỗi sản phẩm chỉ được có một listing trên mỗi website.")
    if any(not item.product_id or not item.source_url for item in items):
        raise ValueError("Mỗi listing phải có product_id và link sản phẩm.")

    xlsx_path = Path(path) if path else _default_path()
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws_master = wb[_MASTER_SHEET]
        ws_listings = wb[_LISTINGS_SHEET]

        master_id = _next_id(ws_master, "master_product_id", "MP", 6)
        sources_display = "; ".join(sorted({it.source.value for it in items}))

        master_header = [c.value for c in ws_master[1]]
        master_row = {
            "master_product_id": master_id,
            "tên_sản_phẩm_chuẩn": canonical_name,
            "số_listing": len(items),
            "số_nguồn": len(items),
            "các_nguồn": sources_display,
            "phương_pháp_ghép": "thủ công (app)",
            "độ_tin_cậy": 1.0,
            "trạng_thái": "thêm thủ công",
        }
        ws_master.append([master_row.get(col) for col in master_header])

        listings_header = [c.value for c in ws_listings[1]]
        for item in items:
            listing_id = _next_id(ws_listings, "listing_id", "L", 7)
            listing_row = {
                "listing_id": listing_id,
                "master_product_id": master_id,
                "source": item.source.value,
                "product_id": item.product_id,
                "source_url": item.source_url,
                "drug_name": canonical_name,
                "loại_bản_ghi": "sản phẩm",
                "phương_pháp_ghép": "thủ công (app)",
                "độ_tin_cậy": 1.0,
                "cần_duyệt": "Không",
            }
            ws_listings.append([listing_row.get(col) for col in listings_header])

        tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(xlsx_path)
    return master_id


def append_or_update_listing(
    master_product_id: str,
    item: CatalogItem,
    canonical_name: str,
    path: str | Path | None = None,
) -> None:
    """Thêm HOẶC SỬA 1 dòng `source_listings` cho 1 site, gắn vào 1
    `master_product_id` ĐÃ CÓ SẴN — khác `append_manual_product` (luôn tạo
    `master_product_id` MỚI). Dùng khi user dán/sửa URL cho 1 site ngay trong bảng
    chi tiết của sản phẩm ĐÃ CÓ trong 'Đã chọn' (xem
    `gui.main_window._on_detail_row_double_click` /
    `CrawlerEngine.set_manual_listing`).

    Tìm dòng cùng (`master_product_id`, `source`): có → SỬA `product_id`/
    `source_url` tại chỗ (không đổi `listing_id`); không có → THÊM dòng mới (và
    cập nhật `số_listing`/`số_nguồn`/`các_nguồn` ở `master_products` cho khớp).
    Cùng cảnh báo hiệu năng/an toàn ghi file như `append_manual_product`."""
    xlsx_path = Path(path) if path else _default_path()
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws_listings = wb[_LISTINGS_SHEET]
        listings_header = [c.value for c in ws_listings[1]]
        h_idx = {col: i + 1 for i, col in enumerate(listings_header)}  # 1-based

        existing_rows: list[int] = []
        for row_num in range(2, ws_listings.max_row + 1):
            mid = ws_listings.cell(row=row_num, column=h_idx["master_product_id"]).value
            src = ws_listings.cell(row=row_num, column=h_idx["source"]).value
            if mid == master_product_id and src == item.source.value:
                existing_rows.append(row_num)

        existing_row = existing_rows[0] if existing_rows else None
        if existing_row is not None:
            ws_listings.cell(
                row=existing_row, column=h_idx["product_id"], value=item.product_id
            )
            ws_listings.cell(
                row=existing_row, column=h_idx["source_url"], value=item.source_url
            )
            if "drug_name" in h_idx:
                ws_listings.cell(
                    row=existing_row, column=h_idx["drug_name"], value=canonical_name
                )
            # Dữ liệu cũ có thể đã vi phạm invariant. Khi người dùng sửa link,
            # giữ đúng một dòng và xóa toàn bộ dòng trùng của cùng master/site.
            for duplicate_row in reversed(existing_rows[1:]):
                ws_listings.delete_rows(duplicate_row, 1)
        else:
            listing_row = {
                "listing_id": _next_id(ws_listings, "listing_id", "L", 7),
                "master_product_id": master_product_id,
                "source": item.source.value,
                "product_id": item.product_id,
                "source_url": item.source_url,
                "drug_name": canonical_name,
                "loại_bản_ghi": "sản phẩm",
                "phương_pháp_ghép": "thủ công (app)",
                "độ_tin_cậy": 1.0,
                "cần_duyệt": "Không",
            }
            ws_listings.append([listing_row.get(col) for col in listings_header])

        _sync_master_row(wb[_MASTER_SHEET], ws_listings, master_product_id)

        tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(xlsx_path)


def _sync_master_row(ws_master, ws_listings, master_product_id: str) -> None:
    """Đồng bộ count/source từ các listing thật sau mỗi lần thêm hoặc sửa."""
    listing_header = [c.value for c in ws_listings[1]]
    listing_idx = {col: i + 1 for i, col in enumerate(listing_header)}
    sources = {
        str(ws_listings.cell(row=row_num, column=listing_idx["source"]).value)
        for row_num in range(2, ws_listings.max_row + 1)
        if ws_listings.cell(
            row=row_num, column=listing_idx["master_product_id"]
        ).value
        == master_product_id
    }
    sources.discard("None")

    header = [c.value for c in ws_master[1]]
    idx = {col: i + 1 for i, col in enumerate(header)}
    id_col = idx["master_product_id"]
    for row_num in range(2, ws_master.max_row + 1):
        if ws_master.cell(row=row_num, column=id_col).value != master_product_id:
            continue
        for count_col in ("số_listing", "số_nguồn"):
            if count_col in idx:
                ws_master.cell(row=row_num, column=idx[count_col], value=len(sources))
        if "các_nguồn" in idx:
            ws_master.cell(
                row=row_num,
                column=idx["các_nguồn"],
                value="; ".join(sorted(sources)),
            )
        return


def rename_master_product(
    master_product_id: str, new_name: str, path: str | Path | None = None
) -> None:
    """Đổi tên chuẩn (`tên_sản_phẩm_chuẩn`) của 1 `master_product_id` — tên này
    dùng CHUNG cho mọi site (bảng tìm/đã chọn nhóm theo tên này), khác
    `drug_name` riêng từng dòng `source_listings`. Gọi từ GUI khi user sửa
    "Tên sản phẩm" trong dialog Sửa (xem
    `gui.main_window._open_edit_listing_dialog`). Không tìm thấy id → không làm
    gì (không raise, để GUI tự quyết định có báo lỗi hay không)."""
    xlsx_path = Path(path) if path else _default_path()
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws_master = wb[_MASTER_SHEET]
        header = [c.value for c in ws_master[1]]
        idx = {col: i + 1 for i, col in enumerate(header)}
        id_col = idx["master_product_id"]
        name_col = idx["tên_sản_phẩm_chuẩn"]
        for row_num in range(2, ws_master.max_row + 1):
            if ws_master.cell(row=row_num, column=id_col).value == master_product_id:
                ws_master.cell(row=row_num, column=name_col, value=new_name)
                break

        tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(xlsx_path)


def delete_listing(
    master_product_id: str, source: SourceName, path: str | Path | None = None
) -> int | None:
    """Xóa 1 listing (1 site) của `master_product_id` khỏi `source_listings` —
    dùng cho nút "Xóa" ở bảng tìm thuốc (chuột phải) và dialog Sửa ở bảng đã
    chọn (xem `gui.main_window._open_edit_listing_dialog`).

    Nếu đó là listing CUỐI CÙNG của sản phẩm, dòng `master_products` cũng bị
    xóa LUÔN (xóa hẳn khỏi catalog — không giữ lại tên rỗng, theo yêu cầu
    "xóa hoàn toàn tên sản phẩm đó khỏi catalog").

    Trả về số listing CÒN LẠI của sản phẩm sau khi xóa (0 = đã xóa hẳn cả sản
    phẩm), hoặc `None` nếu không tìm thấy listing (không có gì bị xóa/ghi)."""
    xlsx_path = Path(path) if path else _default_path()
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws_listings = wb[_LISTINGS_SHEET]
        listings_header = [c.value for c in ws_listings[1]]
        h_idx = {col: i + 1 for i, col in enumerate(listings_header)}

        target_row = None
        for row_num in range(2, ws_listings.max_row + 1):
            mid = ws_listings.cell(row=row_num, column=h_idx["master_product_id"]).value
            src = ws_listings.cell(row=row_num, column=h_idx["source"]).value
            if mid == master_product_id and src == source.value:
                target_row = row_num
                break
        if target_row is None:
            wb.close()
            return None
        ws_listings.delete_rows(target_row)

        remaining_sources: set[str] = set()
        for row_num in range(2, ws_listings.max_row + 1):
            if (
                ws_listings.cell(row=row_num, column=h_idx["master_product_id"]).value
                == master_product_id
            ):
                remaining_sources.add(
                    ws_listings.cell(row=row_num, column=h_idx["source"]).value
                )

        ws_master = wb[_MASTER_SHEET]
        master_header = [c.value for c in ws_master[1]]
        m_idx = {col: i + 1 for i, col in enumerate(master_header)}
        id_col = m_idx["master_product_id"]
        for row_num in range(2, ws_master.max_row + 1):
            if ws_master.cell(row=row_num, column=id_col).value != master_product_id:
                continue
            if remaining_sources:
                if "số_listing" in m_idx:
                    ws_master.cell(
                        row=row_num,
                        column=m_idx["số_listing"],
                        value=len(remaining_sources),
                    )
                if "số_nguồn" in m_idx:
                    ws_master.cell(
                        row=row_num,
                        column=m_idx["số_nguồn"],
                        value=len(remaining_sources),
                    )
                if "các_nguồn" in m_idx:
                    ws_master.cell(
                        row=row_num,
                        column=m_idx["các_nguồn"],
                        value="; ".join(sorted(remaining_sources)),
                    )
            else:
                ws_master.delete_rows(row_num)
            break

        tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(xlsx_path)
    return len(remaining_sources)


def delete_product(
    master_product_id: str, path: str | Path | None = None
) -> int | None:
    """Xóa TOÀN BỘ 1 sản phẩm (mọi listing MỌI site) + dòng `master_products`
    khỏi catalog trong 1 lần ghi — dùng cho nút "Xóa" chuột phải ở bảng tìm
    thuốc (xem `gui.main_window._confirm_delete_product`), KHÁC `delete_listing`
    (chỉ xóa 1 site, có thể vẫn giữ lại các site khác).

    Trả về số listing đã xóa, hoặc `None` nếu không tìm thấy `master_product_id`
    trong `master_products` (không có gì bị xóa/ghi)."""
    xlsx_path = Path(path) if path else _default_path()
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws_master = wb[_MASTER_SHEET]
        master_header = [c.value for c in ws_master[1]]
        m_idx = {col: i + 1 for i, col in enumerate(master_header)}
        id_col = m_idx["master_product_id"]

        master_row = None
        for row_num in range(2, ws_master.max_row + 1):
            if ws_master.cell(row=row_num, column=id_col).value == master_product_id:
                master_row = row_num
                break
        if master_row is None:
            wb.close()
            return None
        ws_master.delete_rows(master_row)

        ws_listings = wb[_LISTINGS_SHEET]
        listings_header = [c.value for c in ws_listings[1]]
        h_idx = {col: i + 1 for i, col in enumerate(listings_header)}
        mid_col = h_idx["master_product_id"]
        rows_to_delete = [
            row_num
            for row_num in range(2, ws_listings.max_row + 1)
            if ws_listings.cell(row=row_num, column=mid_col).value == master_product_id
        ]
        for row_num in reversed(rows_to_delete):
            ws_listings.delete_rows(row_num)

        tmp_path = xlsx_path.with_suffix(".xlsx.tmp")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(xlsx_path)
    return len(rows_to_delete)
