"""Ghi DrugPrice ra Excel (.xlsx) — cùng semantics dedup với CsvWriter.

Ví von: cùng là "bảng giá treo tường" như CSV nhưng bản in màu — header đậm,
dòng giá rẻ nhất mỗi nhóm thuốc tô xanh để nhà thuốc nhìn phát biết ngay
nên nhập nguồn nào.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .csv_writer import CsvWriter, _key
from .models import CSV_HEADERS, DrugPrice

_CHEAPEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_CHEAPEST_FONT = Font(color="006100", bold=True)
_COL_WIDTHS = {
    "drug_name": 38, "canonical_name": 30, "brand": 16, "manufacturer": 22,
    "dosage_form": 14, "strength": 12, "price_vnd": 12, "price_display": 14,
    "source": 16, "source_url": 40, "crawled_at": 20,
}


class ExcelWriter:
    """Merge bản ghi mới vào file .xlsx hiện có, dedup theo (drug_name + source)."""

    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)

    def write(self, prices: list[DrugPrice]) -> int:
        """Trả về tổng số dòng dữ liệu sau khi ghi."""
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

        rows: dict[str, dict[str, str]] = {}

        # 1. Nạp dữ liệu cũ (nếu file đã tồn tại).
        if self.filepath.exists():
            wb_old = load_workbook(self.filepath, read_only=True)
            ws_old = wb_old.active
            headers: list[str] = []
            for i, row in enumerate(ws_old.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else "" for c in row]
                    continue
                d = {h: ("" if v is None else str(v)) for h, v in zip(headers, row)}
                rows[_key(d.get("drug_name", ""), d.get("source", ""))] = d
            wb_old.close()

        # 2. Ghi đè / thêm bản ghi mới (tái dùng chuyển đổi của CsvWriter).
        for price in prices:
            row = CsvWriter._to_row(price)
            rows[_key(row["drug_name"], row["source"])] = row

        # 3. Ghi lại toàn bộ, tô màu dòng rẻ nhất mỗi nhóm.
        wb = Workbook()
        ws = wb.active
        ws.title = "Giá thuốc"
        ws.append(CSV_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for idx, header in enumerate(CSV_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(header, 15)

        ordered = list(rows.values())
        cheapest_keys = _cheapest_keys(ordered)
        for r, d in enumerate(ordered, start=2):
            ws.append([d.get(h, "") for h in CSV_HEADERS])
            if _key(d.get("drug_name", ""), d.get("source", "")) in cheapest_keys:
                for cell in ws[r]:
                    cell.fill = _CHEAPEST_FILL
                    cell.font = _CHEAPEST_FONT

        wb.save(self.filepath)
        return len(rows)


def _cheapest_keys(rows: list[dict[str, str]]) -> set[str]:
    """Key (drug_name+source) của dòng giá thấp nhất trong mỗi nhóm thuốc.

    Nhóm theo canonical_name (fallback drug_name); bỏ qua giá 0 (giá ẩn).
    """
    best: dict[str, tuple[int, str]] = {}  # group -> (price, row_key)
    for d in rows:
        try:
            price = int(float(d.get("price_vnd", "0") or 0))
        except ValueError:
            continue
        if price <= 0:
            continue
        group = (d.get("canonical_name") or d.get("drug_name") or "").strip().lower()
        key = _key(d.get("drug_name", ""), d.get("source", ""))
        if group not in best or price < best[group][0]:
            best[group] = (price, key)
    return {key for _, key in best.values()}


def writer_for(filepath: str | Path) -> CsvWriter | ExcelWriter:
    """Chọn writer theo đuôi file: .xlsx → Excel, còn lại → CSV."""
    if str(filepath).lower().endswith(".xlsx"):
        return ExcelWriter(filepath)
    return CsvWriter(filepath)
