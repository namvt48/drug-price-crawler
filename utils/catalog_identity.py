"""Kiểm tra và làm sạch quan hệ sản phẩm chuẩn ↔ listing theo từng website.

Invariant runtime: mỗi ``(master_product_id, source)`` có tối đa một
``product_id`` và mỗi ``(source, product_id)`` chỉ thuộc một master product.
Các trường hợp không đủ chắc chắn được chuyển sang sheet ``listing_can_duyet``
thay vì chọn bừa một giá có thể thuộc SKU khác.
"""

from __future__ import annotations

import shutil
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
_REVIEW_SHEET = "listing_can_duyet"


@dataclass(frozen=True)
class CatalogIdentityReport:
    original_rows: int
    active_rows: int
    review_rows: int
    auto_selected_pairs: int
    duplicate_pairs_before: int
    duplicate_pairs_after: int
    cross_master_ids_before: int
    cross_master_ids_after: int


@dataclass(frozen=True)
class _Record:
    order: int
    values: tuple[object, ...]
    master_id: str
    source: str
    product_id: str
    listing_name: str


def _count_identity_issues(
    records: list[_Record],
) -> tuple[int, int]:
    by_pair: dict[tuple[str, str], int] = defaultdict(int)
    by_source_id: dict[tuple[str, str], set[str]] = defaultdict(set)
    for record in records:
        by_pair[(record.master_id, record.source)] += 1
        by_source_id[(record.source, record.product_id)].add(record.master_id)
    duplicate_pairs = sum(count > 1 for count in by_pair.values())
    cross_master = sum(len(master_ids) > 1 for master_ids in by_source_id.values())
    return duplicate_pairs, cross_master


def clean_catalog_identity(
    path: str | Path,
    *,
    backup_path: str | Path | None = None,
    preferred_ids: dict[tuple[str, str], str] | None = None,
) -> CatalogIdentityReport:
    """Làm sạch catalog theo ID, không suy luận identity từ tên người dùng đặt."""
    catalog_path = Path(path)
    preferred_ids = preferred_ids or {}
    if backup_path is not None:
        shutil.copy2(catalog_path, Path(backup_path))

    wb = load_workbook(catalog_path, read_only=False, data_only=False)
    try:
        ws_master = wb["master_products"]
        master_header = [cell.value for cell in ws_master[1]]
        master_idx = {name: index for index, name in enumerate(master_header)}
        canonical_names = {
            str(row[master_idx["master_product_id"]]): str(
                row[master_idx["tên_sản_phẩm_chuẩn"]] or ""
            )
            for row in ws_master.iter_rows(min_row=2, values_only=True)
            if row[master_idx["master_product_id"]]
        }

        ws_listings = wb["source_listings"]
        listing_header = [cell.value for cell in ws_listings[1]]
        listing_idx = {name: index for index, name in enumerate(listing_header)}
        records: list[_Record] = []
        for order, row in enumerate(
            ws_listings.iter_rows(min_row=2, values_only=True), start=2
        ):
            master_id = row[listing_idx["master_product_id"]]
            source = row[listing_idx["source"]]
            product_id = row[listing_idx["product_id"]]
            if not master_id or not source or not product_id:
                continue
            records.append(
                _Record(
                    order=order,
                    values=tuple(row),
                    master_id=str(master_id),
                    source=str(source),
                    product_id=str(product_id),
                    listing_name=str(row[listing_idx["drug_name"]] or ""),
                )
            )

        duplicate_before, cross_before = _count_identity_issues(records)
        selected: set[int] = set()
        review: dict[int, str] = {}
        auto_selected = 0

        by_pair: dict[tuple[str, str], list[_Record]] = defaultdict(list)
        for record in records:
            by_pair[(record.master_id, record.source)].append(record)

        for pair_records in by_pair.values():
            if len(pair_records) == 1:
                selected.add(pair_records[0].order)
                continue
            unique_ids = {record.product_id for record in pair_records}
            if len(unique_ids) == 1:
                selected.add(pair_records[0].order)
                for record in pair_records[1:]:
                    review[record.order] = "dòng ID trùng hoàn toàn"
                auto_selected += 1
                continue
            pair_key = (pair_records[0].master_id, pair_records[0].source)
            preferred_id = preferred_ids.get(pair_key)
            winner = next(
                (
                    record
                    for record in pair_records
                    if record.product_id == preferred_id
                ),
                None,
            )
            if winner is not None:
                selected.add(winner.order)
                auto_selected += 1
            for record in pair_records:
                if winner is None or record.order != winner.order:
                    review[record.order] = (
                        f"không phải ID override {preferred_id}"
                        if winner is not None
                        else "nhiều ID cùng site, cần chọn ID rõ ràng"
                    )

        # Một source product ID không được gắn vào nhiều master product.
        by_source_id: dict[tuple[str, str], list[_Record]] = defaultdict(list)
        record_by_order = {record.order: record for record in records}
        for order in selected:
            record = record_by_order[order]
            by_source_id[(record.source, record.product_id)].append(record)
        for collision in by_source_id.values():
            if len({record.master_id for record in collision}) <= 1:
                continue
            explicit = [
                record
                for record in collision
                if preferred_ids.get((record.master_id, record.source))
                == record.product_id
            ]
            winners = {explicit[0].order} if len(explicit) == 1 else set()
            for record in collision:
                if record.order in winners:
                    continue
                selected.discard(record.order)
                review[record.order] = "một product_id đang thuộc nhiều master product"

        active_records = [record for record in records if record.order in selected]
        active_records.sort(key=lambda record: record.order)

        if ws_listings.max_row > 1:
            ws_listings.delete_rows(2, ws_listings.max_row - 1)
        for record in active_records:
            ws_listings.append(record.values)

        if _REVIEW_SHEET in wb.sheetnames:
            del wb[_REVIEW_SHEET]
        ws_review = wb.create_sheet(_REVIEW_SHEET)
        ws_review.append(listing_header + ["tên_sản_phẩm_chuẩn", "lý_do"])
        for record in sorted(records, key=lambda item: item.order):
            if record.order not in review:
                continue
            ws_review.append(
                list(record.values)
                + [canonical_names.get(record.master_id, ""), review[record.order]]
            )

        sources_by_master: dict[str, set[str]] = defaultdict(set)
        for record in active_records:
            sources_by_master[record.master_id].add(record.source)
        for row_num in range(2, ws_master.max_row + 1):
            master_id = str(
                ws_master.cell(
                    row=row_num, column=master_idx["master_product_id"] + 1
                ).value
                or ""
            )
            sources = sorted(sources_by_master.get(master_id, set()))
            for column in ("số_listing", "số_nguồn"):
                if column in master_idx:
                    ws_master.cell(
                        row=row_num, column=master_idx[column] + 1, value=len(sources)
                    )
            if "các_nguồn" in master_idx:
                ws_master.cell(
                    row=row_num,
                    column=master_idx["các_nguồn"] + 1,
                    value="; ".join(sources) or None,
                )
            if not sources and "trạng_thái" in master_idx:
                ws_master.cell(
                    row=row_num,
                    column=master_idx["trạng_thái"] + 1,
                    value="cần duyệt listing",
                )

        duplicate_after, cross_after = _count_identity_issues(active_records)
        tmp_path = catalog_path.with_name(f"{catalog_path.stem}.tmp{catalog_path.suffix}")
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(catalog_path)

    return CatalogIdentityReport(
        original_rows=len(records),
        active_rows=len(active_records),
        review_rows=len(review),
        auto_selected_pairs=auto_selected,
        duplicate_pairs_before=duplicate_before,
        duplicate_pairs_after=duplicate_after,
        cross_master_ids_before=cross_before,
        cross_master_ids_after=cross_after,
    )
