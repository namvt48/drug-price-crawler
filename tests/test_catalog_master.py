"""Tests cho utils.catalog_master.load_master_catalog — dựng file xlsx fixture nhỏ
ngay trong test thay vì phụ thuộc output/catalog_master.xlsx thật
(41k dòng)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils.catalog_master import (
    append_manual_product,
    append_or_update_listing,
    delete_listing,
    delete_product,
    load_master_catalog,
    rename_master_product,
)
from utils.models import CatalogItem, SourceName

_MASTER_HEADER = ["master_product_id", "tên_sản_phẩm_chuẩn"]
_LISTING_HEADER = [
    "listing_id",
    "master_product_id",
    "variant_id",
    "source",
    "product_id",
    "source_url",
    "drug_name",
    "nhà_sản_xuất_xuất_xứ",
]


def _write_workbook(
    path: Path, master_rows: list[tuple], listing_rows: list[tuple]
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws_master = wb.create_sheet("master_products")
    ws_master.append(_MASTER_HEADER)
    for row in master_rows:
        ws_master.append(row)
    ws_listings = wb.create_sheet("source_listings")
    ws_listings.append(_LISTING_HEADER)
    for row in listing_rows:
        ws_listings.append(row)
    wb.save(path)


class TestLoadMasterCatalog:
    def test_missing_file_returns_empty(self, tmp_path: Path) -> None:
        assert load_master_catalog(tmp_path / "nope.xlsx") == []

    def test_basic_load(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Boganic Chuẩn")],
            listing_rows=[
                (
                    "L1",
                    "MP1",
                    "PV1",
                    "Giathuoctot",
                    "p1",
                    "https://a.test/product/p1",
                    "Boganic Forte 100v",
                    "Traphaco",
                ),
                (
                    "L2",
                    "MP1",
                    "PV1",
                    "ChoThuoc247",
                    "2",
                    "https://b.test/san-pham/2.html",
                    "Boganic H100V",
                    "",
                ),
            ],
        )
        items = load_master_catalog(path)
        assert len(items) == 2
        assert {it.product_id for it in items} == {"p1", "2"}
        # Cả 2 item cùng master_product_id phải mang tên CHUẨN (không phải tên gốc listing).
        assert all(it.drug_name == "Boganic Chuẩn" for it in items)
        assert all(it.master_product_id == "MP1" for it in items)
        by_id = {it.product_id: it for it in items}
        assert by_id["p1"].source == SourceName.GIATHUOCTOT
        assert by_id["p1"].manufacturer == "Traphaco"
        assert by_id["p1"].source_url == "https://a.test/product/p1"
        assert by_id["2"].manufacturer == ""
        assert by_id["p1"].search_name == "boganic chuan"
        assert by_id["p1"].source_drug_name == "Boganic Forte 100v"

    def test_duplicate_master_source_is_rejected_without_guessing(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Alaxan (10x10)")],
            listing_rows=[
                (
                    "L1",
                    "MP1",
                    "PV1",
                    "Giathuoctot",
                    "alaxan-10x10",
                    "https://x.test/product/alaxan-10x10",
                    "Alaxan hộp 10 vỉ x 10 viên",
                    "United",
                ),
                (
                    "L2",
                    "MP1",
                    "PV2",
                    "Giathuoctot",
                    "alaxan-25x4",
                    "https://x.test/product/alaxan-25x4",
                    "Alaxan hộp 25 vỉ x 4 viên",
                    "United",
                ),
            ],
        )
        messages: list[str] = []

        items = load_master_catalog(path, log=messages.append)

        assert items == []
        assert any("trùng master/site" in message for message in messages)

    def test_unknown_master_id_falls_back_to_listing_name(self, tmp_path: Path) -> None:
        """master_product_id không có trong master_products (dữ liệu lỗi) — vẫn
        không mất listing, dùng tạm tên gốc của listing."""
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[],
            listing_rows=[
                (
                    "L1",
                    "MP_missing",
                    "PV1",
                    "Giathuoctot",
                    "p1",
                    "https://a.test/product/p1",
                    "Ten Goc",
                    "",
                ),
            ],
        )
        items = load_master_catalog(path)
        assert len(items) == 1
        assert items[0].drug_name == "Ten Goc"

    def test_skips_row_with_invalid_source(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Boganic Chuẩn")],
            listing_rows=[
                ("L1", "MP1", "PV1", "KhongTonTai", "p1", "https://a.test/p1", "X", ""),
                ("L2", "MP1", "PV1", "Giathuoctot", "p2", "https://b.test/product/p2", "Y", ""),
            ],
        )
        items = load_master_catalog(path)
        assert len(items) == 1
        assert items[0].product_id == "p2"

    def test_skips_row_missing_product_id(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Boganic Chuẩn")],
            listing_rows=[
                ("L1", "MP1", "PV1", "Giathuoctot", None, "https://a.test/p1", "X", ""),
            ],
        )
        assert load_master_catalog(path) == []

    def test_skips_row_missing_or_mismatched_product_link(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Boganic Chuẩn")],
            listing_rows=[
                ("L1", "MP1", "PV1", "Giathuoctot", "p1", None, "X", ""),
                (
                    "L2",
                    "MP1",
                    "PV1",
                    "Giathuoctot",
                    "p2",
                    "https://a.test/product/not-p2",
                    "Y",
                    "",
                ),
            ],
        )
        messages: list[str] = []

        assert load_master_catalog(path, log=messages.append) == []
        assert any("link không khớp product_id" in message for message in messages)

    def test_missing_sheets_return_empty(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        wb = Workbook()
        wb.save(path)
        assert load_master_catalog(path) == []

    def test_log_called_on_missing_file(self, tmp_path: Path) -> None:
        messages: list[str] = []
        load_master_catalog(tmp_path / "nope.xlsx", log=messages.append)
        assert any("Không tìm thấy" in m for m in messages)

    def test_load_orders_groups_by_canonical_name_instead_of_source_rows(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_workbook(
            path,
            master_rows=[("MP1", "Alpha"), ("MP2", "Beta")],
            listing_rows=[
                ("L1", "MP2", "PV2", "BachHoaThuoc", "202", "https://x.test/beta--s202", "Beta", ""),
                ("L2", "MP1", "PV1", "BachHoaThuoc", "101", "https://x.test/alpha--s101", "Alpha", ""),
                ("L3", "MP1", "PV1", "ThuocHaPu", "https://x.test/hapu-alpha", "https://x.test/hapu-alpha", "Alpha", ""),
            ],
        )

        items = load_master_catalog(path)

        assert [item.product_id for item in items] == [
            "101",
            "https://x.test/hapu-alpha",
            "202",
        ]


_FULL_MASTER_HEADER = [
    "master_product_id",
    "tên_sản_phẩm_chuẩn",
    "hàm_lượng_chuẩn",
    "hoạt_chất_mô_tả",
    "dạng_bào_chế",
    "nhà_sản_xuất_xuất_xứ",
    "số_listing",
    "số_nguồn",
    "các_nguồn",
    "phương_pháp_ghép",
    "độ_tin_cậy",
    "trạng_thái",
]
_FULL_LISTING_HEADER = [
    "listing_id",
    "master_product_id",
    "variant_id",
    "source",
    "product_id",
    "source_url",
    "drug_name",
    "loại_bản_ghi",
    "tên_sản_phẩm_đã_tách",
    "hàm_lượng",
    "hoạt_chất_mô_tả",
    "dạng_bào_chế",
    "nhà_sản_xuất_xuất_xứ",
    "quy_cách_đóng_gói",
    "pack_signature",
    "phương_pháp_ghép",
    "độ_tin_cậy",
    "cần_duyệt",
]


def _write_full_workbook(
    path: Path, master_rows: list[tuple], listing_rows: list[tuple]
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws_master = wb.create_sheet("master_products")
    ws_master.append(_FULL_MASTER_HEADER)
    for row in master_rows:
        ws_master.append(row)
    ws_listings = wb.create_sheet("source_listings")
    ws_listings.append(_FULL_LISTING_HEADER)
    for row in listing_rows:
        ws_listings.append(row)
    wb.save(path)


def _ci(
    product_id: str, source: SourceName, source_url: str = "https://x.test/p"
) -> CatalogItem:
    return CatalogItem(
        product_id=product_id, drug_name="tmp", source=source, source_url=source_url
    )


class TestAppendManualProduct:
    def test_appends_master_and_listing_rows(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Sản phẩm cũ",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "old-p1",
                    "https://a.test/old",
                    "Sản phẩm cũ",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        items = [
            _ci("new-p1", SourceName.GIATHUOCTOT, "https://a.test/new1"),
            _ci("new-p2", SourceName.CHOTHUOC247, "https://a.test/new2"),
        ]
        new_id = append_manual_product(items, "Sản phẩm mới", path=path)
        assert new_id == "MP000002"

        wb = load_workbook(path, read_only=True)
        master_rows_after = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows_after = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        assert len(master_rows_after) == 3  # header + 1 cũ + 1 mới
        new_master = master_rows_after[-1]
        header = master_rows_after[0]
        idx = {c: i for i, c in enumerate(header)}
        assert new_master[idx["master_product_id"]] == "MP000002"
        assert new_master[idx["tên_sản_phẩm_chuẩn"]] == "Sản phẩm mới"
        assert new_master[idx["số_nguồn"]] == 2
        assert new_master[idx["trạng_thái"]] == "thêm thủ công"

        assert len(listing_rows_after) == 4  # header + 1 cũ + 2 mới
        lidx = {c: i for i, c in enumerate(listing_rows_after[0])}
        assert all(
            r[lidx["master_product_id"]] == "MP000002" for r in listing_rows_after[2:]
        )
        assert {r[lidx["product_id"]] for r in listing_rows_after[2:]} == {
            "new-p1",
            "new-p2",
        }
        assert listing_rows_after[2][lidx["listing_id"]] == "L0000002"
        assert listing_rows_after[3][lidx["listing_id"]] == "L0000003"

    def test_empty_workbook_starts_at_1(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(path, master_rows=[], listing_rows=[])
        new_id = append_manual_product(
            [_ci("p1", SourceName.GIATHUOCTOT)], "Sản phẩm đầu tiên", path=path
        )
        assert new_id == "MP000001"

        wb = load_workbook(path, read_only=True)
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()
        header = listing_rows[0]
        idx = {c: i for i, c in enumerate(header)}
        assert listing_rows[1][idx["listing_id"]] == "L0000001"

    def test_preserves_existing_rows_untouched(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Cũ 1",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "old-p1",
                    "https://a.test/old",
                    "Cũ 1",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        append_manual_product([_ci("new-p1", SourceName.GIATHUOCTOT)], "Mới", path=path)

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        wb.close()
        assert master_rows[1][0] == "MP000001"
        assert master_rows[1][1] == "Cũ 1"
        assert master_rows[1][9] == "entity-resolution"  # phương_pháp_ghép cũ không đổi


class TestAppendOrUpdateListing:
    def test_adds_new_site_bumps_master_counts(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        new_item = _ci("new-p2", SourceName.CHOTHUOC247, "https://b.test/new")
        append_or_update_listing("MP000001", new_item, "Boganic", path=path)

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        assert len(listing_rows) == 3  # header + 1 cũ + 1 mới
        lidx = {c: i for i, c in enumerate(listing_rows[0])}
        assert listing_rows[2][lidx["source"]] == "ChoThuoc247"
        assert listing_rows[2][lidx["product_id"]] == "new-p2"
        assert listing_rows[2][lidx["master_product_id"]] == "MP000001"
        assert listing_rows[2][lidx["listing_id"]] == "L0000002"

        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert master_rows[1][midx["số_listing"]] == 2
        assert master_rows[1][midx["số_nguồn"]] == 2
        assert master_rows[1][midx["các_nguồn"]] == "ChoThuoc247; Giathuoctot"

    def test_updates_existing_site_in_place_no_master_bump(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "old-id",
                    "https://a.test/old",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        updated_item = _ci("new-id", SourceName.GIATHUOCTOT, "https://a.test/updated")
        append_or_update_listing("MP000001", updated_item, "Boganic mới", path=path)

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        assert len(listing_rows) == 2  # header + 1 dòng (SỬA tại chỗ, không thêm dòng)
        lidx = {c: i for i, c in enumerate(listing_rows[0])}
        assert (
            listing_rows[1][lidx["listing_id"]] == "L0000001"
        )  # giữ nguyên listing_id
        assert listing_rows[1][lidx["product_id"]] == "new-id"
        assert listing_rows[1][lidx["source_url"]] == "https://a.test/updated"
        # SỬA dòng đã có PHẢI cập nhật cả drug_name — trước đây chỉ sửa
        # product_id/source_url, bỏ sót drug_name (bug đã sửa cùng feature này).
        assert listing_rows[1][lidx["drug_name"]] == "Boganic mới"

        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert (
            master_rows[1][midx["số_listing"]] == 1
        )  # không tăng vì chỉ SỬA, không thêm site mới

    def test_no_duplicate_source_name_in_cac_nguon(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        # site khác nhưng update lần 2 không được đếm trùng
        append_or_update_listing(
            "MP000001",
            _ci("p2", SourceName.CHOTHUOC247, "https://b.test/p2"),
            "Boganic",
            path=path,
        )
        append_or_update_listing(
            "MP000001",
            _ci("p2-v2", SourceName.CHOTHUOC247, "https://b.test/p2v2"),
            "Boganic",
            path=path,
        )

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert (
            master_rows[1][midx["số_nguồn"]] == 2
        )  # Giathuoctot + ChoThuoc247, không đếm 2 lần
        assert master_rows[1][midx["các_nguồn"]] == "ChoThuoc247; Giathuoctot"
        assert (
            len(listing_rows) == 3
        )  # header + Giathuoctot cũ + ChoThuoc247 (sửa tại chỗ lần 2)


class TestRenameMasterProduct:
    def test_renames_matching_row(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
                (
                    "MP000002",
                    "Panadol",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
            ],
            listing_rows=[],
        )
        rename_master_product("MP000001", "Boganic Forte", path=path)

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        wb.close()
        idx = {c: i for i, c in enumerate(master_rows[0])}
        assert master_rows[1][idx["tên_sản_phẩm_chuẩn"]] == "Boganic Forte"
        assert (
            master_rows[2][idx["tên_sản_phẩm_chuẩn"]] == "Panadol"
        )  # dòng khác không đổi

    def test_unknown_id_is_noop(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[],
        )
        rename_master_product("MP_missing", "Tên mới", path=path)

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        wb.close()
        idx = {c: i for i, c in enumerate(master_rows[0])}
        assert master_rows[1][idx["tên_sản_phẩm_chuẩn"]] == "Boganic"


class TestDeleteListing:
    def test_deletes_one_of_multiple_listings_keeps_master_row(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    2,
                    2,
                    "ChoThuoc247; Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
                (
                    "L0000002",
                    "MP000001",
                    None,
                    "ChoThuoc247",
                    "p2",
                    "https://b.test/p2",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
            ],
        )
        remaining = delete_listing("MP000001", SourceName.GIATHUOCTOT, path=path)
        assert remaining == 1

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        assert len(listing_rows) == 2  # header + ChoThuoc247 còn lại
        lidx = {c: i for i, c in enumerate(listing_rows[0])}
        assert listing_rows[1][lidx["source"]] == "ChoThuoc247"

        assert len(master_rows) == 2  # header + dòng master vẫn còn (chưa hết site)
        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert master_rows[1][midx["số_listing"]] == 1
        assert master_rows[1][midx["số_nguồn"]] == 1
        assert master_rows[1][midx["các_nguồn"]] == "ChoThuoc247"

    def test_deletes_last_listing_removes_master_row_too(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
                (
                    "MP000002",
                    "Panadol",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
                (
                    "L0000002",
                    "MP000002",
                    None,
                    "Giathuoctot",
                    "p2",
                    "https://b.test/p2",
                    "Panadol",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
            ],
        )
        remaining = delete_listing("MP000001", SourceName.GIATHUOCTOT, path=path)
        assert remaining == 0

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        # Sản phẩm bị xóa hoàn toàn khỏi master_products (không giữ tên rỗng).
        assert len(master_rows) == 2  # header + Panadol (MP000002 không bị đụng)
        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert {r[midx["master_product_id"]] for r in master_rows[1:]} == {"MP000002"}

        assert len(listing_rows) == 2  # header + listing của Panadol
        lidx = {c: i for i, c in enumerate(listing_rows[0])}
        assert listing_rows[1][lidx["product_id"]] == "p2"

    def test_not_found_returns_none_and_does_not_modify(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        result = delete_listing("MP000001", SourceName.CHOTHUOC247, path=path)
        assert result is None

        wb = load_workbook(path, read_only=True)
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()
        assert len(listing_rows) == 2  # không mất gì


class TestDeleteProduct:
    def test_deletes_all_listings_and_master_row(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    2,
                    2,
                    "ChoThuoc247; Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
                (
                    "MP000002",
                    "Panadol",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                ),
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
                (
                    "L0000002",
                    "MP000001",
                    None,
                    "ChoThuoc247",
                    "p2",
                    "https://b.test/p2",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
                (
                    "L0000003",
                    "MP000002",
                    None,
                    "Giathuoctot",
                    "p3",
                    "https://c.test/p3",
                    "Panadol",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                ),
            ],
        )
        removed = delete_product("MP000001", path=path)
        assert removed == 2  # 2 listing (Giathuoctot + ChoThuoc247) bị xóa

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()

        # Chỉ còn Panadol (MP000002) — Boganic bị xóa HOÀN TOÀN cả 2 site.
        midx = {c: i for i, c in enumerate(master_rows[0])}
        assert {r[midx["master_product_id"]] for r in master_rows[1:]} == {"MP000002"}

        assert len(listing_rows) == 2  # header + listing của Panadol
        lidx = {c: i for i, c in enumerate(listing_rows[0])}
        assert listing_rows[1][lidx["product_id"]] == "p3"

    def test_not_found_returns_none_and_does_not_modify(self, tmp_path: Path) -> None:
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Boganic",
                    None,
                    None,
                    None,
                    None,
                    1,
                    1,
                    "Giathuoctot",
                    "entity-resolution",
                    0.9,
                    "đã duyệt",
                )
            ],
            listing_rows=[
                (
                    "L0000001",
                    "MP000001",
                    None,
                    "Giathuoctot",
                    "p1",
                    "https://a.test/p1",
                    "Boganic",
                    "sản phẩm",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "entity-resolution",
                    0.9,
                    "Không",
                )
            ],
        )
        result = delete_product("MP_missing", path=path)
        assert result is None

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        listing_rows = list(wb["source_listings"].iter_rows(values_only=True))
        wb.close()
        assert len(master_rows) == 2  # không mất gì
        assert len(listing_rows) == 2  # không mất gì

    def test_master_row_with_no_listings_still_removed(self, tmp_path: Path) -> None:
        """Sản phẩm có dòng master_products nhưng KHÔNG có listing nào (vd
        'chưa có id' từ import gốc) — vẫn xóa được dòng master, trả về 0."""
        path = tmp_path / "catalog.xlsx"
        _write_full_workbook(
            path,
            master_rows=[
                (
                    "MP000001",
                    "Chưa có id",
                    None,
                    None,
                    None,
                    None,
                    0,
                    0,
                    "",
                    "chưa tìm được",
                    0,
                    "chưa có site listing",
                )
            ],
            listing_rows=[],
        )
        removed = delete_product("MP000001", path=path)
        assert removed == 0

        wb = load_workbook(path, read_only=True)
        master_rows = list(wb["master_products"].iter_rows(values_only=True))
        wb.close()
        assert len(master_rows) == 1  # chỉ còn header
