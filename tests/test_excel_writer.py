"""Tests cho utils.excel_writer — ghi .xlsx, merge dedup, tô màu giá rẻ nhất."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from utils.csv_writer import CsvWriter
from utils.excel_writer import ExcelWriter, writer_for
from utils.models import CSV_HEADERS, DrugPrice, SourceName


def _dp(
    name: str,
    source: SourceName = SourceName.GIATHUOCTOT,
    price: int = 1000,
    canonical: str = "",
) -> DrugPrice:
    return DrugPrice(
        drug_name=name,
        canonical_name=canonical or name,
        price_vnd=price,
        price_display=f"{price:,}đ",
        source=source,
    )


def _rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(c) for c in next(it)]
    out = [{h: ("" if v is None else str(v)) for h, v in zip(headers, row)} for row in it]
    wb.close()
    return out


class TestExcelWrite:
    def test_write_creates_file_with_headers(self, tmp_path: Path) -> None:
        f = tmp_path / "out.xlsx"
        total = ExcelWriter(f).write([_dp("A"), _dp("B")])
        assert total == 2
        assert f.exists()
        wb = load_workbook(f)
        ws = wb.active
        assert [c.value for c in ws[1]] == CSV_HEADERS
        assert ws.freeze_panes == "A2"
        assert ws[1][0].font.bold is True
        wb.close()

    def test_merge_dedups_by_name_and_source(self, tmp_path: Path) -> None:
        f = tmp_path / "out.xlsx"
        w = ExcelWriter(f)
        w.write([_dp("A", price=1000)])
        total = w.write([_dp("A", price=2000), _dp("B")])  # A bị ghi đè giá mới
        assert total == 2
        rows = _rows(f)
        a = next(r for r in rows if r["drug_name"] == "A")
        assert a["price_vnd"] == "2000"

    def test_cheapest_row_highlighted(self, tmp_path: Path) -> None:
        f = tmp_path / "out.xlsx"
        ExcelWriter(f).write([
            _dp("Boganic X", SourceName.GIATHUOCTOT, 2000, canonical="Boganic"),
            _dp("Boganic Y", SourceName.THUOCSI, 1500, canonical="Boganic"),
        ])
        wb = load_workbook(f)
        ws = wb.active
        fills = {}
        for row in ws.iter_rows(min_row=2):
            name = row[0].value
            fills[name] = row[0].fill.start_color.rgb
        wb.close()
        assert fills["Boganic Y"] == "00C6EFCE"
        assert fills["Boganic X"] != "00C6EFCE"

    def test_zero_price_never_highlighted(self, tmp_path: Path) -> None:
        f = tmp_path / "out.xlsx"
        ExcelWriter(f).write([_dp("A", price=0, canonical="A")])
        wb = load_workbook(f)
        ws = wb.active
        assert ws[2][0].fill.start_color.rgb != "00C6EFCE"
        wb.close()


class TestWriterFor:
    def test_xlsx_gets_excel_writer(self, tmp_path: Path) -> None:
        assert isinstance(writer_for(tmp_path / "x.XLSX"), ExcelWriter)

    def test_csv_gets_csv_writer(self, tmp_path: Path) -> None:
        assert isinstance(writer_for(tmp_path / "x.csv"), CsvWriter)
        assert isinstance(writer_for(tmp_path / "x.txt"), CsvWriter)
