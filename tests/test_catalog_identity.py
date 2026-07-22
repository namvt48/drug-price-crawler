from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils.catalog_identity import clean_catalog_identity


def _write_catalog(
    path: Path,
    masters: list[tuple[str, str]],
    listings: list[tuple[str, str, str, str, str]],
) -> None:
    wb = Workbook()
    ws_master = wb.active
    ws_master.title = "master_products"
    ws_master.append(
        [
            "master_product_id",
            "tên_sản_phẩm_chuẩn",
            "số_listing",
            "số_nguồn",
            "các_nguồn",
            "trạng_thái",
        ]
    )
    for master_id, name in masters:
        ws_master.append([master_id, name, 0, 0, "", "đã ghép"])

    ws_listing = wb.create_sheet("source_listings")
    ws_listing.append(
        [
            "listing_id",
            "master_product_id",
            "source",
            "product_id",
            "source_url",
            "drug_name",
        ]
    )
    for listing_id, master_id, source, product_id, drug_name in listings:
        ws_listing.append(
            [listing_id, master_id, source, product_id, f"https://x/{product_id}", drug_name]
        )
    wb.save(path)


def _active_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["source_listings"].iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def test_cleanup_uses_explicit_id_override_not_name_similarity(tmp_path: Path) -> None:
    path = tmp_path / "catalog.xlsx"
    _write_catalog(
        path,
        [("MP1", "Alaxan (10x10)")],
        [
            ("L1", "MP1", "Giathuoctot", "alaxan-10x10", "Alaxan hộp 10 vỉ x 10 viên"),
            ("L2", "MP1", "Giathuoctot", "alaxan-25x4", "Alaxan hộp 25 vỉ x 4 viên"),
        ],
    )

    report = clean_catalog_identity(
        path, preferred_ids={("MP1", "Giathuoctot"): "alaxan-10x10"}
    )

    assert report.duplicate_pairs_before == 1
    assert report.duplicate_pairs_after == 0
    assert [row[3] for row in _active_rows(path)] == ["alaxan-10x10"]


def test_cleanup_moves_ambiguous_seller_duplicates_to_review(tmp_path: Path) -> None:
    path = tmp_path / "catalog.xlsx"
    _write_catalog(
        path,
        [("MP1", "Boganic Forte (H/50v)")],
        [
            ("L1", "MP1", "ThuocSi", "seller-a", "Boganic Forte (H/50v)"),
            ("L2", "MP1", "ThuocSi", "seller-b", "Boganic Forte (H/50v)"),
        ],
    )

    report = clean_catalog_identity(path)

    assert report.review_rows == 2
    assert _active_rows(path) == []
    wb = load_workbook(path, read_only=True, data_only=True)
    assert wb["listing_can_duyet"].max_row == 3
    wb.close()


def test_cleanup_resolves_source_id_assigned_to_two_masters(tmp_path: Path) -> None:
    path = tmp_path / "catalog.xlsx"
    _write_catalog(
        path,
        [("MP1", "Biotin 5mg (20 viên)"), ("MP2", "Biotin HD (100 viên)")],
        [
            ("L1", "MP1", "ThuocTot3Mien", "646", "Biotin 5mg hộp 2 vỉ x 10 viên"),
            ("L2", "MP2", "ThuocTot3Mien", "646", "Biotin 5mg hộp 2 vỉ x 10 viên"),
        ],
    )

    report = clean_catalog_identity(
        path, preferred_ids={("MP1", "ThuocTot3Mien"): "646"}
    )

    assert report.cross_master_ids_before == 1
    assert report.cross_master_ids_after == 0
    assert [(row[1], row[3]) for row in _active_rows(path)] == [("MP1", "646")]


def test_cleanup_keeps_single_listing_regardless_of_mutable_display_name(
    tmp_path: Path,
) -> None:
    path = tmp_path / "catalog.xlsx"
    _write_catalog(
        path,
        [("MP1", "Biotin (20 viên)")],
        [
            (
                "L1",
                "MP1",
                "ThuocHaPu",
                "biotin-hd-5x20",
                "Biotin HD hộp 5 vỉ x 20 viên",
            )
        ],
    )

    report = clean_catalog_identity(path)

    assert report.review_rows == 0
    assert [row[3] for row in _active_rows(path)] == ["biotin-hd-5x20"]
